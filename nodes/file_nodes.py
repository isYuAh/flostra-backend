from __future__ import annotations

import json
import mimetypes
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

import httpx

from runtime_files import (
    build_file_ref,
    download_url_bytes,
    from_base64,
    get_file_record,
    is_file_ref,
    new_file_id,
    sha256_hex,
    to_base64,
    upsert_file_record,
)
from io_utils import FileIoService

from .base import JsonDict, WorkflowNode, register_node


def _guess_mime(name: Optional[str], fallback: str = "application/octet-stream") -> str:
    if not name:
        return fallback
    mime, _ = mimetypes.guess_type(name)
    return mime or fallback


def _max_bytes(params: JsonDict, default: int = 8_000_000) -> int:
    try:
        val = int(params.get("max_bytes") or default)
        return max(1, val)
    except Exception:
        return default


@register_node
class FileDownloadNode(WorkflowNode):
    """从 URL 下载文件并生成 FileRef；可选 lazy 仅创建引用，等用到再下载。"""

    type = "file.download"

    @classmethod
    def get_schema(cls) -> JsonDict:
        return {
            "type": cls.type,
            "label": "下载文件",
            "description": "从 URL 下载文件，输出 File 引用（默认写入 cacheMap，惰性读取可选）",
            "category": "文件",
            "icon": "download",
            "inputs": [
                {"name": "url", "label": "URL", "type": "string", "required": True},
                {"name": "name", "label": "文件名(可选)", "type": "string", "required": False},
                {"name": "mime", "label": "MIME(可选)", "type": "string", "required": False},
            ],
            "outputs": [
                {"name": "file", "label": "文件", "type": "object", "required": True},
            ],
            "parameters": [
                {
                    "name": "lazy",
                    "label": "惰性下载（只创建引用）",
                    "type": "boolean",
                    "widget": "switch",
                    "default": False,
                },
                {
                    "name": "max_bytes",
                    "label": "最大下载大小（字节）",
                    "type": "number",
                    "widget": "input",
                    "default": 8_000_000,
                },
            ],
        }

    @classmethod
    async def run(cls, inputs: JsonDict, params: JsonDict, context: JsonDict = None) -> JsonDict:
        if not context:
             raise ValueError("文件节点需要运行时上下文")
        url = inputs.get("url")
        if not isinstance(url, str) or not url:
            raise ValueError("file.download 缺少 url")
        
        # Simple validation
        try:
             parsed = urlparse(url)
             if parsed.scheme not in {"http", "https"}:
                 raise ValueError("仅支持 http/https URL")
        except Exception:
             raise ValueError("url 非法")

        name = inputs.get("name") if isinstance(inputs.get("name"), str) else None
        mime = inputs.get("mime") if isinstance(inputs.get("mime"), str) else None
        lazy = bool(params.get("lazy", False))
        max_bytes = _max_bytes(params)

        if not name:
            try:
                path = urlparse(url).path
                if path and path != "/":
                    name = path.rsplit("/", 1)[-1] or None
            except Exception:
                name = None

        resolved_mime = mime or _guess_mime(name)
        file_id = new_file_id()

        if lazy:
            upsert_file_record(context, file_id, url=url, name=name or "", mime=resolved_mime, source="url")
            return {"file": build_file_ref(file_id=file_id, name=name, mime=resolved_mime, source="url")}

        data = await download_url_bytes(url, max_bytes=max_bytes)
        upsert_file_record(
            context,
            file_id,
            bytes=data,
            url=url,
            name=name or "",
            mime=resolved_mime,
            size=len(data),
            sha256=sha256_hex(data),
            source="url",
        )
        return {
            "file": build_file_ref(
                file_id=file_id,
                name=name,
                mime=resolved_mime,
                size=len(data),
                sha256=sha256_hex(data),
                source="url",
            )
        }


@register_node
class FileToObjectNode(WorkflowNode):
    """把 FileRef 展开为可序列化对象（不包含文件内容）。"""

    type = "file.to_object"

    @classmethod
    def get_schema(cls) -> JsonDict:
        return {
            "type": cls.type,
            "label": "文件转对象",
            "description": "将 File 引用展开为元数据对象（不含内容）",
            "category": "文件",
            "icon": "file",
            "inputs": [
                {"name": "file", "label": "文件", "type": "object", "required": True},
            ],
            "outputs": [
                {"name": "object", "label": "对象", "type": "object", "required": True},
            ],
            "parameters": [
                {
                    "name": "include",
                    "label": "包含字段",
                    "type": "array",
                    "widget": "input",
                    "required": False,
                    "desc": "可选：仅输出这些字段（例如 id,name,mime,size,sha256,source）。为空则输出默认元数据。",
                }
            ],
        }

    @classmethod
    async def run(cls, inputs: JsonDict, params: JsonDict, context: JsonDict = None) -> JsonDict:
        if not context:
             raise ValueError("文件节点需要运行时上下文")
        file_obj = inputs.get("file")
        if not is_file_ref(file_obj):
            raise ValueError("file.to_object 需要输入 FileRef（{__kind:'file', id:...}）")

        file_id = str(file_obj.get("id"))
        rec = get_file_record(context, file_id) or {}

        meta: Dict[str, Any] = {
            "id": file_id,
            "name": file_obj.get("name") or rec.get("name") or None,
            "mime": file_obj.get("mime") or rec.get("mime") or None,
            "size": file_obj.get("size") or rec.get("size") or (len(rec["bytes"]) if isinstance(rec.get("bytes"), (bytes, bytearray)) else None),
            "sha256": file_obj.get("sha256") or rec.get("sha256") or None,
            "source": file_obj.get("source") or rec.get("source") or None,
            "url": rec.get("url"),
        }

        include = params.get("include")
        if isinstance(include, list) and include:
            picked: Dict[str, Any] = {}
            for k in include:
                if not isinstance(k, str):
                    continue
                if k in meta:
                    picked[k] = meta.get(k)
            meta = picked

        # 丢掉 None，保持输出干净
        meta = {k: v for k, v in meta.items() if v is not None}
        return {"object": meta}


@register_node
class FileReadBase64Node(WorkflowNode):
    """将 FileRef 内容读取为 base64（显式拉取内容，慎用）。"""

    type = "file.read_base64"

    @classmethod
    def get_schema(cls) -> JsonDict:
        return {
            "type": cls.type,
            "label": "读取 Base64",
            "description": "读取文件内容为 Base64（会占用内存与事件体积，建议仅小文件使用）",
            "category": "文件",
            "icon": "binary",
            "inputs": [
                {"name": "file", "label": "文件", "type": "object", "required": True},
            ],
            "outputs": [
                {"name": "base64", "label": "Base64", "type": "string", "required": True},
                {"name": "mime", "label": "MIME", "type": "string", "required": False},
                {"name": "name", "label": "文件名", "type": "string", "required": False},
            ],
            "parameters": [
                {
                    "name": "max_bytes",
                    "label": "最大读取大小（字节）",
                    "type": "number",
                    "widget": "input",
                    "default": 2_000_000,
                }
            ],
        }

    @classmethod
    async def run(cls, inputs: JsonDict, params: JsonDict, context: JsonDict = None) -> JsonDict:
        if not context:
             raise ValueError("文件节点需要运行时上下文")
        file_obj = inputs.get("file")
        if not is_file_ref(file_obj):
            raise ValueError("file.read_base64 需要输入 FileRef")
        # max_bytes = _max_bytes(params, default=2_000_000) # IO Service currently doesn't limit size in read, TODO
        
        data, name, mime = await FileIoService.get_file_bytes(context, file_obj)
        
        return {
            "base64": to_base64(data),
            "mime": mime or file_obj.get("mime") or None,
            "name": name or file_obj.get("name") or None,
        }


@register_node
class FileParseNode(WorkflowNode):
    """解析文件为文本/行等结构（仅内置轻量解析；复杂格式依赖可选包）。"""

    type = "file.parse"

    @classmethod
    def get_schema(cls) -> JsonDict:
        return {
            "type": cls.type,
            "label": "解析文件",
            "description": "将文件解析为文本或表格行（csv 内置；docx/xlsx 需可选依赖）",
            "category": "文件",
            "icon": "parse",
            "inputs": [
                {"name": "file", "label": "文件", "type": "object", "required": True},
            ],
            "outputs": [
                {"name": "text", "label": "文本", "type": "string", "required": False},
                {"name": "rows", "label": "行", "type": "array", "required": False},
                {"name": "meta", "label": "元数据", "type": "object", "required": False},
            ],
            "parameters": [
                {"name": "format", "label": "格式", "type": "string", "widget": "select", "required": False, "extra": {"options": [
                    {"label": "auto", "value": "auto"},
                    {"label": "text", "value": "text"},
                    {"label": "json", "value": "json"},
                    {"label": "csv", "value": "csv"},
                    {"label": "docx", "value": "docx"},
                    {"label": "xlsx", "value": "xlsx"},
                ]}, "default": "auto"},
                {"name": "max_bytes", "label": "最大读取大小（字节）", "type": "number", "widget": "input", "default": 8_000_000},
                {"name": "text_max_chars", "label": "文本截断上限", "type": "number", "widget": "input", "default": 20_000},
                {"name": "rows_max", "label": "最大行数", "type": "number", "widget": "input", "default": 200},
                {"name": "sheet", "label": "xlsx sheet 索引", "type": "number", "widget": "input", "default": 0},
                {"name": "encoding", "label": "文本编码", "type": "string", "widget": "input", "default": "utf-8"},
            ],
        }

    @classmethod
    async def run(cls, inputs: JsonDict, params: JsonDict, context: JsonDict = None) -> JsonDict:
        if not context:
             raise ValueError("文件节点需要运行时上下文")
        file_obj = inputs.get("file")
        if not is_file_ref(file_obj):
            raise ValueError("file.parse 需要输入 FileRef")

        fmt = (params.get("format") or "auto").lower()
        max_bytes = _max_bytes(params, default=8_000_000)
        text_max_chars = int(params.get("text_max_chars") or 20_000)
        rows_max = int(params.get("rows_max") or 200)
        sheet = int(params.get("sheet") or 0)
        encoding = str(params.get("encoding") or "utf-8")

        # Use IO Service to get content
        data, name, mime = await FileIoService.get_file_bytes(context, file_obj)
        
        if not name:
             name = str(file_obj.get("name") or "")
        if not mime:
             mime = str(file_obj.get("mime") or "")

        if fmt == "auto":
            ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
            if mime.startswith("text/") or ext in {"txt", "md", "log"}:
                fmt = "text"
            elif ext in {"json"}:
                fmt = "json"
            elif ext in {"csv"}:
                fmt = "csv"
            elif ext in {"docx"}:
                fmt = "docx"
            elif ext in {"xlsx"}:
                fmt = "xlsx"
            else:
                fmt = "text"

        meta = {"name": name or None, "mime": mime or None, "size": len(data)}

        if fmt == "text":
            text = data.decode(encoding, errors="replace")
            if len(text) > text_max_chars:
                text = text[:text_max_chars]
                meta["truncated"] = True
            return {"text": text, "meta": meta}

        if fmt == "json":
            text = data.decode(encoding, errors="replace")
            try:
                obj = json.loads(text)
            except Exception as exc:  # noqa: BLE001
                raise ValueError(f"JSON 解析失败：{exc}") from exc
            return {"meta": meta, "text": json.dumps(obj, ensure_ascii=False)}

        if fmt == "csv":
            import csv
            import io

            text = data.decode(encoding, errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows: List[List[Any]] = []
            for i, row in enumerate(reader):
                if i >= rows_max:
                    meta["truncated"] = True
                    break
                rows.append(list(row))
            return {"rows": rows, "meta": meta}

        if fmt == "docx":
            try:
                import docx  # type: ignore
            except ImportError:
                raise ValueError("缺少依赖 python-docx；请安装后重试")
            import io

            doc = docx.Document(io.BytesIO(data))
            text = "\n".join([p.text for p in doc.paragraphs])
            if len(text) > text_max_chars:
                text = text[:text_max_chars]
                meta["truncated"] = True
            return {"text": text, "meta": meta}

        if fmt == "xlsx":
            try:
                import openpyxl  # type: ignore
            except ImportError:
                raise ValueError("缺少依赖 openpyxl；请安装后重试")
            import io

            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheets = wb.sheetnames
            if sheet < 0 or sheet >= len(sheets):
                raise ValueError(f"sheet 索引超出范围: {sheet}")
            ws = wb[sheets[sheet]]
            rows: List[List[Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= rows_max:
                    meta["truncated"] = True
                    break
                rows.append([cell for cell in row])
            meta["sheet"] = sheets[sheet]
            return {"rows": rows, "meta": meta}

        raise ValueError(f"不支持的 format: {fmt!r}")


@register_node
class FileGenerateNode(WorkflowNode):
    """从文本/JSON 生成文件（写入 cacheMap）。"""

    type = "file.generate"

    @classmethod
    def get_schema(cls) -> JsonDict:
        return {
            "type": cls.type,
            "label": "生成文件",
            "description": "将文本/JSON/CSV 内容生成临时文件引用（写入 cacheMap）",
            "category": "文件",
            "icon": "file-plus",
            "inputs": [
                {"name": "content", "label": "内容", "type": ["string", "object", "array"], "required": True},
                {"name": "name", "label": "文件名", "type": "string", "required": False},
                {"name": "mime", "label": "MIME(可选)", "type": "string", "required": False},
            ],
            "outputs": [
                {"name": "file", "label": "文件", "type": "object", "required": True},
            ],
            "parameters": [
                {"name": "format", "label": "输出格式", "type": "string", "widget": "select", "extra": {"options": [
                    {"label": "text", "value": "text"},
                    {"label": "json", "value": "json"},
                    {"label": "csv(rows)", "value": "csv"},
                ]}, "default": "text"},
                {"name": "encoding", "label": "编码", "type": "string", "widget": "input", "default": "utf-8"},
            ],
        }

    @classmethod
    async def run(cls, inputs: JsonDict, params: JsonDict, context: JsonDict = None) -> JsonDict:
        if not context:
             raise ValueError("文件节点需要运行时上下文")
        ctx = context # Alias
        content = inputs.get("content")
        name = inputs.get("name") if isinstance(inputs.get("name"), str) else None
        mime = inputs.get("mime") if isinstance(inputs.get("mime"), str) else None
        fmt = (params.get("format") or "text").lower()
        encoding = str(params.get("encoding") or "utf-8")

        if fmt == "json":
            text = json.dumps(content, ensure_ascii=False, indent=2)
            if not name:
                name = "data.json"
            resolved_mime = mime or "application/json"
            data = text.encode(encoding)
        elif fmt == "csv":
            import csv
            import io

            if not isinstance(content, list):
                raise ValueError("csv 格式要求 content 为 rows 数组（二维数组）")
            buf = io.StringIO()
            writer = csv.writer(buf)
            for row in content:
                if isinstance(row, list):
                    writer.writerow(row)
                else:
                    writer.writerow([row])
            if not name:
                name = "data.csv"
            resolved_mime = mime or "text/csv"
            data = buf.getvalue().encode(encoding)
        else:
            text = content if isinstance(content, str) else json.dumps(content, ensure_ascii=False)
            if not name:
                name = "data.txt"
            resolved_mime = mime or _guess_mime(name, fallback="text/plain")
            data = str(text).encode(encoding)

        file_id = new_file_id()
        upsert_file_record(
            ctx,
            file_id,
            bytes=data,
            name=name,
            mime=resolved_mime,
            size=len(data),
            sha256=sha256_hex(data),
            source="generated",
        )
        return {
            "file": build_file_ref(
                file_id=file_id,
                name=name,
                mime=resolved_mime,
                size=len(data),
                sha256=sha256_hex(data),
                source="generated",
            )
        }
