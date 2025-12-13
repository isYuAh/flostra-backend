from __future__ import annotations

from abc import ABC, abstractmethod
import base64
import io
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Type


@dataclass
class LLMContext:
    """
    LLM 节点运行时上下文。
    插件通过修改此对象来影响最终的 LLM 调用和结果。
    """

    # 基础元数据
    session_id: Optional[str] = None
    user_query: str = ""

    # 输入参数（可被插件修改）
    system_prompt: str = ""
    messages: List[Dict[str, Any]] = field(default_factory=list)

    # 工具相关
    tools: List[Dict[str, Any]] = field(default_factory=list)  # OpenAI tool schema 列表
    tool_functions: Dict[str, Callable[..., Any]] = field(default_factory=dict)  # name -> callable

    # 执行结果
    ai_response_text: str = ""
    ai_raw_response: Any = None

    # 额外透传/调试数据
    extra_data: Dict[str, Any] = field(default_factory=dict)


class WorkflowPlugin(ABC):
    """所有插件的基类"""

    type: str  # 唯一标识

    def __init__(self, config: Dict[str, Any]):
        self.config = config or {}

    @classmethod
    @abstractmethod
    def get_schema(cls) -> Dict[str, Any]:
        """返回插件的配置 schema（前端用）"""

    async def on_pre_process(self, context: LLMContext) -> None:
        """LLM 调用前的钩子"""
        return None

    async def on_post_process(self, context: LLMContext) -> None:
        """LLM 调用后的钩子"""
        return None


_PLUGIN_REGISTRY: Dict[str, Type[WorkflowPlugin]] = {}


def register_plugin(cls: Type[WorkflowPlugin]) -> Type[WorkflowPlugin]:
    p_type = getattr(cls, "type", None)
    if not p_type:
        raise ValueError("Plugin class must define a non-empty 'type'")
    if p_type in _PLUGIN_REGISTRY:
        raise ValueError(f"Duplicated plugin type: {p_type}")
    _PLUGIN_REGISTRY[p_type] = cls
    return cls


def get_plugin_class(plugin_type: str) -> Optional[Type[WorkflowPlugin]]:
    return _PLUGIN_REGISTRY.get(plugin_type)


def list_plugins_schema() -> List[Dict[str, Any]]:
    """给前端：返回所有插件的 schema 列表"""
    return [cls.get_schema() for cls in _PLUGIN_REGISTRY.values()]


class PluginExecutor:
    """串行执行插件生命周期"""

    def __init__(self, plugins_conf: List[Dict[str, Any]], context: LLMContext):
        self.context = context
        self.plugins: List[WorkflowPlugin] = []
        for conf in plugins_conf or []:
            p_type = conf.get("type")
            p_cls = get_plugin_class(p_type)
            if p_cls:
                self.plugins.append(p_cls(conf.get("config", {})))

    async def run_pre_process(self) -> None:
        for plugin in self.plugins:
            await plugin.on_pre_process(self.context)

    async def run_post_process(self) -> None:
        for plugin in self.plugins:
            await plugin.on_post_process(self.context)


@register_plugin
class LanguageEnforceZhPlugin(WorkflowPlugin):
    """强制输出中文的系统提示附加"""

    type = "lang-zh"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "语言：中文",
            "description": "在 system prompt 中加入“请用中文回答”",
            "kind": "plugin",
            "parameters": [],
        }

    async def on_pre_process(self, context: LLMContext) -> None:
        enforce = "请用中文回答，用简洁、准确的中文输出。"
        if context.system_prompt:
            context.system_prompt = context.system_prompt + "\n" + enforce
        else:
            context.system_prompt = enforce


@register_plugin
class InputTruncatePlugin(WorkflowPlugin):
    """截断用户输入，防止超长 prompt 导致超时/超费"""

    type = "input-truncate"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "输入截断",
            "description": "限制用户输入长度，超过则截断并追加提示",
            "kind": "plugin",
            "parameters": [
                {
                    "name": "max_chars",
                    "label": "最大字符数",
                    "type": "number",
                    "widget": "input",
                    "default": 4000,
                    "required": True,
                    "extra": {"min": 1, "placeholder": "默认 4000"},
                },
                {
                    "name": "notice",
                    "label": "截断提示",
                    "type": "string",
                    "widget": "textarea",
                    "default": "（用户输入过长，已截断）",
                    "required": False,
                },
            ],
        }

    async def on_pre_process(self, context: LLMContext) -> None:
        max_chars = int(self.config.get("max_chars") or 4000)
        notice = self.config.get("notice") or "（用户输入过长，已截断）"
        if not context.messages:
            return
        last = context.messages[-1]
        if not isinstance(last, dict):
            return
        content = str(last.get("content", ""))
        if len(content) > max_chars:
            truncated = content[:max_chars]
            last["content"] = truncated + "\n" + notice


@register_plugin
class DatetimeToolPlugin(WorkflowPlugin):
    """提供当前时间工具，便于模型生成带时间的回答"""

    type = "tool-datetime"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "工具：当前时间",
            "description": "向模型暴露一个获取当前时间的工具 now()",
            "kind": "tool",
            "parameters": [],
        }

    async def on_pre_process(self, context: LLMContext) -> None:
        import datetime

        context.tools.append(
            {
                "type": "function",
                "function": {
                    "name": "now",
                    "description": "获取当前时间（ISO 8601，UTC）",
                    "parameters": {"type": "object", "properties": {}, "required": []},
                },
            }
        )

        def _now() -> str:
            return datetime.datetime.utcnow().isoformat() + "Z"

        context.tool_functions["now"] = _now


@register_plugin
class SafetyFilterPlugin(WorkflowPlugin):
    """输出前敏感词过滤（命中则替换为掩码）"""

    type = "safety-filter"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "安全过滤",
            "description": "对模型输出应用敏感词替换（纯文本级别）",
            "kind": "plugin",
            "parameters": [
                {
                    "name": "blocked_words",
                    "label": "敏感词列表",
                    "type": "array",
                    "widget": "json-editor",
                    "required": True,
                    "default": ["傻逼", "操", "fuck"],
                    "extra": {"placeholder": "[\"word1\", \"word2\"]"},
                },
                {
                    "name": "mask",
                    "label": "替换为",
                    "type": "string",
                    "widget": "input",
                    "default": "***",
                    "required": False,
                },
                {
                    "name": "case_insensitive",
                    "label": "大小写不敏感",
                    "type": "boolean",
                    "widget": "switch",
                    "default": True,
                    "required": False,
                },
            ],
        }

    async def on_post_process(self, context: LLMContext) -> None:
        text = context.ai_response_text or ""
        words = self.config.get("blocked_words") or []
        mask = self.config.get("mask") or "***"
        ci = bool(self.config.get("case_insensitive", True))
        if not text or not words:
            return
        for w in words:
            if not w:
                continue
            flags = re.IGNORECASE if ci else 0
            text = re.sub(re.escape(str(w)), mask, text, flags=flags)
        context.ai_response_text = text
        # raw 也尽量同步
        if isinstance(context.ai_raw_response, dict):
            try:
                choices = context.ai_raw_response.get("choices")
                if choices and isinstance(choices, list):
                    msg = choices[0].get("message", {})
                    if isinstance(msg, dict) and msg.get("content"):
                        msg["content"] = text
            except Exception:
                pass


@register_plugin
class HttpToolPlugin(WorkflowPlugin):
    """提供受限域名的 HTTP 请求工具"""

    type = "tool-http"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "工具：HTTP",
            "description": "向模型暴露 http_get/http_post（限制域名白名单）",
            "kind": "tool",
            "parameters": [
                {
                    "name": "allowed_hosts",
                    "label": "允许的主机名",
                    "type": "array",
                    "widget": "json-editor",
                    "required": True,
                    "default": ["jsonplaceholder.typicode.com"],
                    "extra": {"placeholder": "[\"api.example.com\"]"},
                },
                {
                    "name": "timeout",
                    "label": "超时（秒）",
                    "type": "number",
                    "widget": "input",
                    "default": 8,
                    "required": False,
                },
            ],
        }

    async def on_pre_process(self, context: LLMContext) -> None:
        import httpx
        import urllib.parse

        allowed = set(self.config.get("allowed_hosts") or [])
        timeout = float(self.config.get("timeout") or 8)

        def _check_host(url: str) -> None:
            host = urllib.parse.urlparse(url).hostname or ""
            if host not in allowed:
                raise ValueError(f"host not allowed: {host}")

        async def _http_get(url: str, headers: Optional[Dict[str, str]] = None, params: Optional[Dict[str, Any]] = None) -> Any:
            _check_host(url)
            async with httpx.AsyncClient(timeout=timeout) as client:
                r = await client.get(url, headers=headers, params=params)
                try:
                    return r.json()
                except Exception:
                    return r.text

        async def _http_post(url: str, json_body: Optional[Any] = None, headers: Optional[Dict[str, str]] = None) -> Any:
            _check_host(url)
            async with httpx.AsyncClient(timeout=timeout) as client:
                r = await client.post(url, json=json_body, headers=headers)
                try:
                    return r.json()
                except Exception:
                    return r.text

        # 注册 tools schema
        context.tools.extend(
            [
                {
                    "type": "function",
                    "function": {
                        "name": "http_get",
                        "description": "GET 请求（受限域名）",
                        "parameters": {
                            "type": "object",
                            "properties": {
                                "url": {"type": "string"},
                                "headers": {"type": "object"},
                                "params": {"type": "object"},
                            },
                            "required": ["url"],
                        },
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "http_post",
                        "description": "POST 请求（受限域名，JSON body）",
                        "parameters": {
                            "type": "object",
                            "properties": {
                                "url": {"type": "string"},
                                "json_body": {"type": "object"},
                                "headers": {"type": "object"},
                            },
                            "required": ["url"],
                        },
                    },
                },
            ]
        )

        context.tool_functions["http_get"] = _http_get
        context.tool_functions["http_post"] = _http_post


@register_plugin
class MathToolPlugin(WorkflowPlugin):
    """提供安全的四则运算工具"""

    type = "tool-math"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "工具：计算器",
            "description": "暴露 calc(expression) 安全计算四则运算",
            "kind": "tool",
            "parameters": [],
        }

    async def on_pre_process(self, context: LLMContext) -> None:
        import ast
        import operator

        ops = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.Pow: operator.pow,
            ast.USub: operator.neg,
        }

        def _eval_expr(expr: str) -> float:
            """
            安全计算简单表达式，只允许数字和加减乘除乘方及括号。
            """
            node = ast.parse(expr, mode="eval").body

            def _eval(n):
                if isinstance(n, ast.Num):  # type: ignore[attr-defined]
                    return n.n
                if isinstance(n, ast.BinOp) and type(n.op) in ops:
                    return ops[type(n.op)](_eval(n.left), _eval(n.right))
                if isinstance(n, ast.UnaryOp) and type(n.op) in ops:
                    return ops[type(n.op)](_eval(n.operand))
                raise ValueError("unsupported expression")

            return float(_eval(node))

        context.tools.append(
            {
                "type": "function",
                "function": {
                    "name": "calc",
                    "description": "计算四则运算表达式（如 1+2*3）",
                    "parameters": {
                        "type": "object",
                        "properties": {"expression": {"type": "string"}},
                        "required": ["expression"],
                    },
                },
            }
        )
        context.tool_functions["calc"] = _eval_expr


@register_plugin
class AttachmentToolPlugin(WorkflowPlugin):
    """基于附件列表提供受控文件访问/解析工具"""

    type = "tool-attachments"

    @classmethod
    def get_schema(cls) -> Dict[str, Any]:
        return {
            "type": cls.type,
            "label": "工具：附件",
            "description": "为模型暴露 list_files/get_base64/docx_to_text/xlsx_to_rows，数据来源于当前对话附带的 attachments",
            "kind": "tool",
            "parameters": [
                {
                    "name": "max_bytes",
                    "label": "单文件大小上限（字节）",
                    "type": "number",
                    "widget": "input",
                    "default": 8_000_000,
                    "required": False,
                },
                {
                    "name": "text_max_chars",
                    "label": "docx 文本截断上限",
                    "type": "number",
                    "widget": "input",
                    "default": 20_000,
                    "required": False,
                },
                {
                    "name": "rows_max",
                    "label": "xlsx 最大行数",
                    "type": "number",
                    "widget": "input",
                    "default": 200,
                    "required": False,
                },
            ],
        }

    async def on_pre_process(self, context: LLMContext) -> None:
        attachments = context.extra_data.get("attachments") if isinstance(context.extra_data, dict) else None
        if not attachments:
            return

        max_bytes = int(self.config.get("max_bytes") or 8_000_000)
        text_max_chars = int(self.config.get("text_max_chars") or 20_000)
        rows_max = int(self.config.get("rows_max") or 200)

        def _est_size_b64(b64_str: str) -> int:
            length = len(b64_str or "")
            # 估算实际字节大小，考虑 padding 略微放宽
            return int(length * 0.75)

        def _find_att(name: str) -> Optional[Dict[str, Any]]:
            for att in attachments:
                if isinstance(att, dict) and att.get("name") == name:
                    return att
            return None

        def _get_att_data(name: str) -> Dict[str, Any]:
            att = _find_att(name)
            if not att:
                raise ValueError(f"附件未找到: {name}")
            b64 = att.get("content_base64") or ""
            size = _est_size_b64(b64)
            if size > max_bytes:
                raise ValueError(f"附件过大（约 {size} bytes > {max_bytes} 上限）")
            return {"name": att.get("name") or "", "mime": att.get("mime") or "", "b64": b64, "size": size}

        def _list_files() -> List[Dict[str, Any]]:
            result: List[Dict[str, Any]] = []
            for att in attachments:
                if not isinstance(att, dict):
                    continue
                b64 = att.get("content_base64") or ""
                result.append(
                    {
                        "name": att.get("name") or "",
                        "mime": att.get("mime") or "",
                        "size": _est_size_b64(b64),
                    }
                )
            return result

        def _get_base64(name: str) -> Dict[str, Any]:
            info = _get_att_data(name)
            return {"name": info["name"], "mime": info["mime"], "base64": info["b64"]}

        def _docx_to_text(name: str, max_chars: Optional[int] = None) -> Dict[str, Any]:
            info = _get_att_data(name)
            limit = min(int(max_chars or text_max_chars), text_max_chars)
            try:
                import docx  # type: ignore
            except ImportError:
                return {"error": "missing dependency python-docx；请安装后重试"}
            try:
                data = base64.b64decode(info["b64"])
                doc = docx.Document(io.BytesIO(data))
                text = "\n".join([p.text for p in doc.paragraphs])
                if len(text) > limit:
                    text = text[:limit]
                return {"name": info["name"], "mime": info["mime"], "text": text, "truncated": len(text) >= limit}
            except Exception as exc:  # noqa: BLE001
                return {"error": f"docx 解析失败: {exc}"}

        def _xlsx_to_rows(name: str, sheet: int = 0, max_rows: Optional[int] = None) -> Dict[str, Any]:
            info = _get_att_data(name)
            limit = min(int(max_rows or rows_max), rows_max)
            try:
                import openpyxl  # type: ignore
            except ImportError:
                return {"error": "missing dependency openpyxl；请安装后重试"}
            try:
                data = base64.b64decode(info["b64"])
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                sheets = wb.sheetnames
                idx = sheet if isinstance(sheet, int) else 0
                if idx < 0 or idx >= len(sheets):
                    return {"error": f"sheet 索引超出范围: {sheet}"}
                ws = wb[sheets[idx]]
                rows: List[List[Any]] = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= limit:
                        break
                    rows.append([cell for cell in row])
                return {"name": info["name"], "mime": info["mime"], "sheet": sheets[idx], "rows": rows}
            except Exception as exc:  # noqa: BLE001
                return {"error": f"xlsx 解析失败: {exc}"}

        # 注册 tools schema
        context.tools.extend(
            [
                {
                    "type": "function",
                    "function": {
                        "name": "list_files",
                        "description": "列出当前对话可用的附件元数据",
                        "parameters": {"type": "object", "properties": {}, "required": []},
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "get_base64",
                        "description": "获取指定附件的 Base64 内容",
                        "parameters": {
                            "type": "object",
                            "properties": {"name": {"type": "string"}},
                            "required": ["name"],
                        },
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "docx_to_text",
                        "description": "解析 docx 为纯文本（自动截断）",
                        "parameters": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "max_chars": {"type": "number"},
                            },
                            "required": ["name"],
                        },
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "xlsx_to_rows",
                        "description": "读取 xlsx 指定 sheet 的前若干行",
                        "parameters": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "sheet": {"type": "number"},
                                "max_rows": {"type": "number"},
                            },
                            "required": ["name"],
                        },
                    },
                },
            ]
        )

        context.tool_functions["list_files"] = _list_files
        context.tool_functions["get_base64"] = _get_base64
        context.tool_functions["docx_to_text"] = _docx_to_text
        context.tool_functions["xlsx_to_rows"] = _xlsx_to_rows
